from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import cv2
import os
import re
from PIL import Image as Img
from io import BytesIO
from makedir import make_dir

wb = Workbook()
ws = wb.active

ws.title = "Extract_Image"
ws_new = wb.create_sheet(title="Fail_Image")

## 시트 색 변경
# ws.sheet_properties.tabColor = "헥스코드"
ws.sheet_properties.tabColor = "00ff00"
ws_new.sheet_properties.tabColor = "890400"


'''
#사용법
ws.column_dimensions[열인덱스(알파벳)].width = 열너비값
ws.row_dimensions[행번호].height = 행너비값

열너비는 0~255포인트(0~2045픽셀), 행높이는 0~409포인트(0~546픽셀)의 범위에서 지정가능
1포인트 == 0.035cm
'''


## GFPGAN 추출, 미추출 이미지 + WEBP 제외
def img_match(true_folder, true_list, cmp_list):
    t_image = {}
    for g in true_list:
    
        ## 이미지 포맷 == WEBP일 경우, openpyxl save가 안됨 - 변환
        path = f"{true_folder}\\{g}"
        img = Img.open(path)

        if img.format == "WEBP":
            print(f"{g} 이미지 포맷: {img.format}")
            convert_img = img.convert("RGB")
            convert_img.save(path, 'jpeg')

            
        img_text = re.sub(r'[.].+', '', g)
        
        
        for cmp in cmp_list:
            cmp_text = re.sub(r'[_].+', '', cmp)
            
            
            if img_text == cmp_text:
                t_image.setdefault(g,[]).append(cmp)
                # t_image.append(g)
            ## 한 이미지에서 2개 이상 Fac detection했을때
            # t_image.setdefault(x[0],[]).append(num)

    ## ws_new = wb.create_sheet(title="Fail_Image")
    f_image = []
    for x in true_list:
        ## <desktop.ini> 파일 존재
        if x == "desktop.ini":
            continue
        
        if x not in t_image.keys():
            f_image.append(x)
            
    return t_image, f_image

## 문자가 섞여있는 리스트에서 숫자를 기준으로 sorting
def digit_sort(n):
    n = int(re.findall("\d+", n)[0])
    return n

## openCV 경로 한글일때 = image.shape확인시
def korean_path(path, img):
    try:
        origin_img = cv2.imread(f"{path}\\{img}")
        return origin_img.shape
    
    ## 한글
    except:
        img_path = f"{path}\\{img}"
        with open(img_path, 'rb') as f:
            data = f.read()
            
        data_io = BytesIO(data)
        image = Img.open(data_io)
        
        return image.size
        

## 경로 == r + "path"
def to_excel(true_folder, cmp_folder, res_folder, save_name):
    
    make_dir("excel_sheet")
    
    ## 이미지 불러오기
    true_list = os.listdir(true_folder)
    cmp_list = os.listdir(cmp_folder)
    
    ## GFPGAN 추출, 미추출 확인
    t_img, f_img = img_match(true_folder, true_list, cmp_list)

    ## cell 시작 위치
    start_row = 2
    for i in range(len(t_img)):
        ## Image(이미지파일경로)
        key_img = list(t_img.keys())[i]
        true_img = Image(f"{true_folder}\\{key_img}")
        
        ## 한글 경로일겨우
        size = korean_path(true_folder, key_img)
        
        ## Pixel 
        true_img.width = 500
        true_img.height = 500

        start_num = i + start_row
        ws.add_image(true_img, f"A{start_num}")
        
        for v in range(len(t_img[key_img])):

            cmp_img = Image(f"{cmp_folder}\\{t_img[key_img][v]}")
            res = Image(f"{res_folder}\\{t_img[key_img][v]}")
            
            ws.add_image(cmp_img, f"C{start_num+v}")
            ws.add_image(res, f"E{start_num+v}")

        ## 한 이미지에서 2개 이상의 Face detect
        if len(t_img[key_img]) >= 2:
            start_row += len(t_img[key_img])-1
        
        ## 없을때
        if not t_img[key_img]:
            continue
        
        ## 원본 이미지 크기와 파일명
        ws.cell(row=start_num, column=1).value = f"이미지 크기: {size[1]}X{size[0]}, 파일명: {t_img[key_img]}"
        ws.cell(row= start_num, column= 1).font = Font(size=20)
        
    
    ## 추출 이미지 개수만큼 cell 조정
    for i in range(len(cmp_list)):
        ## cell 이미지 크기에 맞춤
        ws.row_dimensions[i+2].height = 450
    
    
    ## 미추출 이미지
    for i in range(len(f_img)):
   
        false_img = Image(f"{true_folder}\\{f_img[i]}")
        size = korean_path(true_folder, f_img[i])
        
        ## Pixel 
        false_img.width = 500
        false_img.height = 500
        
        
        ws_new.cell(row=i+2, column=1).value = f"이미지 크기: ({size[1]}, {size[0]}), 파일명:  {f_img[i]}"
        ws_new.cell(row=i+2, column=1).font = Font(size=20)
        ws_new.add_image(false_img, f"A{i+2}")
        
        ws_new.row_dimensions[i+2].height = 450

    ## -----ws-----
    ws.cell(row= 1, column= 1).value = f"원본, {len(t_img)}개"
    ws.cell(row= 1, column= 1).font = Font(size=20)
    
    ws.cell(row= 1, column= 3).value = "비교 1024X512"
    ws.cell(row= 1, column= 3).font = Font(size=20)
    
    ws.cell(row= 1, column= 5).value = f"추출 이미지512X512, {len(cmp_list)}개"
    ws.cell(row= 1, column= 5).font = Font(size=20)
    
    ## 추출 이미지 cell 사이즈
    ws.column_dimensions['A'].width = 64
    ws.column_dimensions['C'].width = 127
    ws.column_dimensions['E'].width = 64
    
    ## -----ws_new-----
    ws_new.column_dimensions['A'].width = 64 
    ws_new.cell(row= 1, column= 1).value = f"원본, {len(f_img)}개"
    ws_new.cell(row= 1, column= 1).font = Font(size=20)
    
    ## 이미지 엑셀에 저장한 다음 파일 이동해야함
    # - WebP : 구글에서 만든 이미지 파일 포맷 / 웹 고속화를 위해 개발된 새로운 압축 포맷
    wb.save(f"excel_sheet\\{save_name}.xlsx")
    
    ## 미추출 이미지 파일 이동(=false_image 디렉토리)
    for m in range(len(f_img)):
        img_file = f"{true_folder}\\{f_img[m]}" # .jpg
        print(img_file)
    #     ## 파일명 변경 후 파일 디렉토리로 이동
    #     os.rename(img_file, f"false_image\\{true_folder}_{f_img[m]}")


## to_excel(스크래핑 폴더 경로, GFPGAN CMP 경로, GFPGAN restored 경로, 저장할 폴더명)
to_excel(r"scraping_folder\여성 의류",
        r"compare\여성의류cmp\cmp",
        r"compare\여성의류restored\restored_faces",
        "여성의류")
    